"""
Excel View Parser (Steps 1-3)

Reads an xlsx file, extracts structured data from a specified tab,
and maps cell fill colors to status values.

Usage:
    from parser import parse_workbook, parse_sheet
    
    # Get all tabs
    tabs = parse_workbook("data/sot.xlsx")
    
    # Parse a specific tab
    data = parse_sheet("data/sot.xlsx", "MSA Premium Detail")
"""

import openpyxl
from typing import Optional


# --- Status source (active) ---
# In the real SOT, status lives as TEXT in this column on MSA Premium Detail.
# Values: Parity, In Progress, Gap, By Design.
STATUS_COLUMN = "MCA Web MSA Gap"

# Separator rows (visual dividers in the sheet, not real features). Skipped.
# Matched against the first column, case-insensitive.
SEPARATOR_MARKERS = {"LINE BREAK"}

# --- Color mapping (STAGED, not active yet) ---
# Urmila wants to move the status source from text to cell color soon, once the
# SOT is fully color-coded and consistent. See NOW.md. The color meaning is
# column-specific (the same hex means different things in different columns):
#   Column F (MCA Web MSA Gap): light green = Parity, yellow = In Progress,
#                               orange = Gap, blue = By Design
#   Columns G/H (Timeline):     orange (FFC000) = In Progress
# Exact hexes to be finalized once the teammate finishes coloring. Known so far:
#   E2EFDA = light green (Parity in column F)
#   FFC000 = orange (In Progress in Timeline columns)
COLOR_STATUS_MAP = {
    "E2EFDA": "Parity",   # column F light green (partial today)
    "FFC000": "In Progress",  # Timeline orange
    "no_fill": "Gap",
}

# Default tolerance for color matching (0-255 per channel)
COLOR_TOLERANCE = 30


def parse_workbook(filepath: str) -> list[str]:
    """Return the list of tab (sheet) names in the workbook."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def parse_sheet(filepath: str, sheet_name: str = "MSA Premium Detail") -> list[dict]:
    """
    Parse a single sheet from the workbook into a list of row dicts.

    Each row dict has:
    - One key per column header, with the (whitespace-stripped) cell value
    - A normalized "status" key, read from the text column STATUS_COLUMN

    Trailing columns with no header are dropped. Empty rows are skipped.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    headers = _read_headers(ws)

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        all_empty = True

        for col_idx, header in enumerate(headers):
            value = ws.cell(row=row_idx, column=col_idx + 1).value
            # Strip whitespace on text values so queries match reliably
            if isinstance(value, str):
                value = value.strip()
            if value is not None and value != "":
                all_empty = False
            row_data[header] = value

        # Skip rows where every cell is empty
        if all_empty:
            continue

        # Skip separator rows (e.g. "LINE BREAK" dividers), not real features
        first_value = row_data.get(headers[0])
        if isinstance(first_value, str) and first_value.upper() in SEPARATOR_MARKERS:
            continue

        # Normalized status, read from the text column (active source)
        row_data["status"] = row_data.get(STATUS_COLUMN)

        rows.append(row_data)

    wb.close()
    return rows


def _read_headers(ws) -> list[str]:
    """
    Read row-1 headers. Drops trailing columns that have no header (blank spacer
    columns). Any remaining blank header gets a positional name.
    """
    raw = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        raw.append(str(val).strip() if val is not None else None)
    # Trim trailing empty-header columns
    while raw and raw[-1] is None:
        raw.pop()
    # Name any remaining blank header by position
    return [h if h is not None else f"Column_{i + 1}" for i, h in enumerate(raw)]


def get_schema(filepath: str, sheet_name: str = "MSA Premium Detail") -> dict:
    """
    Return the schema of a sheet: column names, which are color-status columns,
    and the color mapping in use.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    
    ws = wb[sheet_name]

    headers = _read_headers(ws)

    wb.close()

    return {
        "sheet_name": sheet_name,
        "columns": headers,
        "status_source": "text",
        "status_column": STATUS_COLUMN,
        "row_count": None,  # Filled after parsing
    }


def _extract_fill_color(cell) -> str:
    """
    Extract the background fill color from a cell.
    Returns the hex color string (e.g., "92D050") or "no_fill".
    """
    fill = cell.fill
    
    if fill is None or fill.fill_type is None or fill.fill_type == "none":
        return "no_fill"
    
    if fill.start_color and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb)
        # openpyxl sometimes returns "00000000" for no fill
        if rgb == "00000000":
            return "no_fill"
        # Strip alpha channel if present (AARRGGBB -> RRGGBB)
        if len(rgb) == 8:
            rgb = rgb[2:]
        return rgb.upper()
    
    return "no_fill"


def _map_color_to_status(color_hex: str) -> str:
    """
    Map a hex color to a status string.
    Uses exact match first, then falls back to nearest color within tolerance.
    """
    if color_hex == "no_fill":
        return COLOR_STATUS_MAP["no_fill"]
    
    # Exact match
    if color_hex in COLOR_STATUS_MAP:
        return COLOR_STATUS_MAP[color_hex]
    
    # Nearest match within tolerance
    best_match = None
    best_distance = float("inf")
    
    for known_hex, status in COLOR_STATUS_MAP.items():
        if known_hex == "no_fill":
            continue
        distance = _color_distance(color_hex, known_hex)
        if distance < best_distance:
            best_distance = distance
            best_match = status
    
    if best_distance <= COLOR_TOLERANCE * 3:  # Sum of RGB channel tolerances
        return best_match
    
    # No match found
    return f"Unknown ({color_hex})"


def _color_distance(hex1: str, hex2: str) -> int:
    """Calculate Manhattan distance between two hex colors."""
    try:
        r1, g1, b1 = int(hex1[0:2], 16), int(hex1[2:4], 16), int(hex1[4:6], 16)
        r2, g2, b2 = int(hex2[0:2], 16), int(hex2[2:4], 16), int(hex2[4:6], 16)
        return abs(r1 - r2) + abs(g1 - g2) + abs(b1 - b2)
    except (ValueError, IndexError):
        return float("inf")


# --- Quick test when run directly ---
if __name__ == "__main__":
    import sys
    
    filepath = sys.argv[1] if len(sys.argv) > 1 else "data/test_sot.xlsx"
    
    print("=" * 60)
    print(f"FILE: {filepath}")
    print("=" * 60)
    
    # Step 1: List tabs
    tabs = parse_workbook(filepath)
    print(f"\nTabs: {tabs}")
    
    # Step 1: Show schema
    schema = get_schema(filepath)
    print(f"\nColumns: {schema['columns']}")
    print(f"Status source: {schema['status_source']} (column: {schema['status_column']})")

    # Steps 2-3: Parse data
    data = parse_sheet(filepath)
    print(f"\nRows parsed: {len(data)}")

    # Show first 5 rows
    print(f"\nFirst 5 rows:")
    for row in data[:5]:
        feature = str(row.get("Feature", "") or "")
        priority = str(row.get("Priority", "") or "")
        dri = str(row.get("DRI", "") or "")
        status = str(row.get("status", "") or "")
        print(f"  {priority:<4} | {feature:<40} | {dri:<20} | {status}")

    # Show status distribution
    from collections import Counter
    counts = Counter(r.get("status") for r in data)
    print(f"\nStatus distribution:")
    for status, count in counts.most_common():
        print(f"  {status}: {count}")
